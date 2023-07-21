import concurrent.futures
from PyQt5.Qt import *
from PyQt5 import uic
import sys
import urllib.request
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class My_Ui(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = uic.loadUi("主窗口.ui", self)
        self.ui.select.clicked.connect(self.selectFileDialog)
        self.ui.start.clicked.connect(self.folder_create)

    def selectFileDialog(self):
        result = QFileDialog.getOpenFileName(self, "选择一个excel文件", "./")
        self.ui.FileName.setText(result[0])

    def downloadMessage(self, flag):  # 关于下载的弹窗提示
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("提示")
        if flag == 1:
            msg.setText("已开始下载")
        elif flag == 0:
            msg.setText("下载完成")
        elif flag == 2:
            msg.setIcon(QMessageBox.Warning)
            msg.setText("该文件中不包含图片链接，请重新选择")
        elif flag == 3:
            msg.setIcon(QMessageBox.Warning)
            msg.setText("文件格式错误")
        msg.exec()

    def folder_create(self):  # 文件夹创建
        save_folder = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') + "\图片"
        try:
            os.makedirs(save_folder, exist_ok=False)
            self.get_url(save_folder)
        except OSError as e:
            msg = QMessageBox()
            msg.setWindowTitle("提示")
            msg.setIcon(QMessageBox.Question)
            msg.setText("图片文件夹已存在")
            msg.setInformativeText("是否删除？")
            msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            result = msg.exec()
            if result == QMessageBox.Ok:
                shutil.rmtree(save_folder)

    def get_url(self, save_folder):  # 获取下载url
        excelPath = self.ui.FileName.text()
        sheet_name = "已签到"
        try:
            workbook = load_workbook(excelPath)
            sheet = workbook[sheet_name]
            link_list = []
            # 获取下载链接
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        # cell.value = cell.hyperlink.target
                        link_list.append(cell.hyperlink.target)
            if link_list:
                self.downLoad(link_list, save_folder)
            else:
                self.downloadMessage(2)
        except InvalidFileException:
            self.downloadMessage(3)

    def download_img(self, url, save_path):  # 图片下载器
        urllib.request.urlretrieve(url, save_path)

    def downLoad(self, link_list, save_folder):  # 多线程下载
        self.downloadMessage(1)
        self.ui.start.setEnabled(False)
        with concurrent.futures.ThreadPoolExecutor() as executor:
            futures = []
            for i, url in enumerate(link_list):
                filename = f"{i}.jpg"
                sava_path = os.path.join(save_folder, filename)
                future = executor.submit(self.download_img, url, sava_path)
                futures.append(future)

            # for future in concurrent.futures.as_completed(futures):
            #     try:
            #         future.result()
            #     except Exception as e:
            #         print(f"下载出错：{e}")
        self.downloadMessage(0)
        self.ui.start.setEnabled(True)


app = QApplication(sys.argv)
my = My_Ui()
my.show()
sys.exit(app.exec_())
